from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QWidget, QVBoxLayout, QLabel, QPushButton, QFileDialog, \
    QHeaderView, QHBoxLayout
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor
import openpyxl
import pandas as pd
from data_manager import DataManager

class ExcelTableScreen(QWidget):
    def __init__(self, data_manager: DataManager):
        super().__init__()
        self.data_manager = data_manager

        self.point_df = pd.read_excel('./dane/sklep_1.xlsx' ,sheet_name='punkty', usecols='B:F', skiprows=0,
                                      nrows=100, keep_default_na=False)
        self.data_manager.set_data("points", self.point_df)

        self.map_df = pd.read_excel('./dane/sklep_1.xlsx', sheet_name='mapa', usecols='B:BE', skiprows=0,
                                    nrows=29)
        self.data_manager.set_data("map", self.map_df)

        self.layout = QVBoxLayout()

        # Initially set visibility to False
        self.label = QLabel("Zaimportowane dane (Ostanie 2 kolumny zostaną wygnerowane automatycznie)")
        self.label.setStyleSheet("font-size: 15px; color: white;")
        self.label.setVisible(False)
        self.layout.addWidget(self.label, alignment=Qt.AlignmentFlag.AlignHCenter)

        self.table_widget = QTableWidget()
        stylesheet = "::section{Background-color:rgb(69, 67, 84);}"
        self.table_widget.horizontalHeader().setStyleSheet(stylesheet)

        stylesheet = "::section{Background-color:rgb(69, 67, 84);}"
        self.table_widget.verticalHeader().setStyleSheet(stylesheet)

        self.table_widget.setStyleSheet("background-color: rgb(69, 67, 84); color: white;border:2px;border-style: none;")

        self.layout.addWidget(self.table_widget)

        self.punkty_file_label = QLabel("Plik z mapą: ")
        self.punkty_file_label.setVisible(False)
        self.punkty_file_label.setStyleSheet("font-size: 15px; color: white;")
        self.layout.addWidget(self.punkty_file_label)

        # Add a new QHBoxLayout for buttons
        button_layout = QHBoxLayout()

        self.load_point_button = QPushButton(self)
        self.load_point_button.setStyleSheet("image: url(./grafika/but_wczytaj.png);"
                                            "width: 120px;"
                                            "height: 40px;"
                                            "background-color: transparent;")
        self.load_point_button.clicked.connect(self.load_point_data)
        button_layout.addWidget(self.load_point_button)

        self.save_button = QPushButton(self)
        self.save_button.setStyleSheet("image: url(./grafika/but_zapisz.png);"
                                       "width: 120px;"
                                       "height: 40px;"
                                       "background-color: transparent;")
        self.save_button.clicked.connect(self.save_changes)
        button_layout.addWidget(self.save_button)

        self.back_button = QPushButton(self)
        self.back_button.setStyleSheet("image: url(./grafika/but_powrot.png);"
                                       "width: 120px;"
                                       "height: 40px;"
                                       "background-color: transparent;")
        self.back_button.clicked.connect(self.go_back)
        button_layout.addWidget(self.back_button)

        # Add the button layout to the main layout
        self.layout.addLayout(button_layout)

        self.setLayout(self.layout)
        self.punkty_file_path = None

        self.point_df = None
        self.map_df = None

        self.setAutoFillBackground(True)
        p = self.palette()
        p.setColor(self.backgroundRole(), QColor(69, 67, 84))  # You can set any color you want here
        self.setPalette(p)

    def load_point_data(self, fromSave=False):
        try:
            file_dialog = QFileDialog()
            if not fromSave:
                file_path, _ = file_dialog.getOpenFileName(self, "Open Excel File (Punkty)", "", "Excel Files (*.xlsx)")
                self.punkty_file_path = file_path
            else:
                file_path = self.punkty_file_path

            if file_path:
                workbook = openpyxl.load_workbook(file_path)

                # Assuming you have two sheets named 'Points' and 'Map' in the Excel file
                point_sheet = workbook['punkty']
                map_sheet = workbook['mapa']

                # Clear the table before loading new data
                self.table_widget.clear()

                # Load Points data
                self.table_widget.setRowCount(point_sheet.max_row)
                self.table_widget.setColumnCount(point_sheet.max_column)

                data = []
                for row_index, row in enumerate(point_sheet.iter_rows(values_only=True)):
                    row_data = []
                    for col_index, value in enumerate(row):
                        if value is not None:
                            item = QTableWidgetItem(str(value))
                            self.table_widget.setItem(row_index, col_index, item)
                            row_data.append(str(value))
                    if row_data:  # Skip empty rows
                        data.append(row_data)

                self.point_df = pd.read_excel(self.punkty_file_path, sheet_name='punkty', usecols='B:F', skiprows=0,
                                              nrows=100, keep_default_na=False)
                self.data_manager.set_data("points", self.point_df)

                self.map_df = pd.read_excel(self.punkty_file_path, sheet_name='mapa', usecols='B:BE', skiprows=0,
                                            nrows=29)
                self.data_manager.set_data("map", self.map_df)

                # Update the label with the loaded file name
                self.punkty_file_label.setText(f"Wczytany plik: {file_path}")
                self.punkty_file_label.setStyleSheet("font-size: 15px; color: white;")

                # Set visibility to True after loading
                self.label.setVisible(True)
                self.punkty_file_label.setVisible(True)

        except Exception as e:
            print(f"Error loading Punkty Excel data: {e}")

    def save_changes(self):
        try:
            if self.punkty_file_path:
                workbook = openpyxl.load_workbook(self.punkty_file_path)
                sheet = workbook['punkty']

                for row in range(self.table_widget.rowCount()):
                    for col in range(self.table_widget.columnCount()):
                        item = self.table_widget.item(row, col)
                        if item is not None:
                            sheet.cell(row=row + 1, column=col + 1, value=item.text())

                workbook.save(self.punkty_file_path)
                print(f"Changes saved successfully for Punkty File: {self.punkty_file_path}")

                # Update the DataFrame after saving changes
                self.load_point_data(True)

        except Exception as e:
            print(f"Error saving changes to Punkty Excel: {e}")

    def go_back(self):
        self.parent().setCurrentIndex(0)
