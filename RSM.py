from PyQt6.QtWidgets import QVBoxLayout, QLabel, QPushButton, QWidget, QCheckBox, QHBoxLayout, QTableView, QButtonGroup, QSizePolicy
from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QFormLayout, QDialog, QHeaderView, QLineEdit, QFileDialog, QMessageBox
from PyQt6.QtCore import Qt, QAbstractTableModel
from PyQt6.QtGui import QColor
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import tabulate
from matplotlib.ticker import (MultipleLocator, AutoMinorLocator)
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

import alg_RSM
from data_manager import DataManager
import a_star


class MatplotlibWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.canvas = FigureCanvas(plt.figure())
        layout = QVBoxLayout()
        layout.addWidget(self.canvas)
        self.canvas.figure.patch.set_facecolor((0.27, 0.26, 0.33))
        self.setLayout(layout)


class KerfusTableModel(QAbstractTableModel):
    def __init__(self, kerfus_tab, parent=None):
        super().__init__(parent)
        self.data = kerfus_tab

    def rowCount(self, parent):
        return len(self.data)

    def columnCount(self, parent):
        return len(self.data[0])

    def data(self, index, role):
        if role == Qt.ItemDataRole.DisplayRole:
            value = self.data[index.row()][index.column()]
            return str(value) if value is not None else ""  # Set empty string for None values
        return None



class ScreenRSM(QWidget):
    def __init__(self, data_manager: DataManager):
        super().__init__()
        self.data_manager = data_manager
        self.data_manager.set_data("user_classes", None)
        self.data_manager.set_data("selection", [0, 1])

        self.layout = QHBoxLayout()  # Use QHBoxLayout instead of QVBoxLayout

        right_layout = QVBoxLayout()
        left_layout = QVBoxLayout()


        self.topsis_label = QLabel(self)
        self.topsis_label.setText("RSM")
        self.topsis_label.setStyleSheet("font-size: 100px; color: white;")
        self.topsis_label.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        left_layout.addWidget(self.topsis_label)

        self.execute_button = QPushButton(self)
        self.execute_button.setStyleSheet("image: url(./grafika/but_oblicz.png);"
                                         "width: 120px;"
                                         "height: 60px;"
                                         "margin: 0px;"
                                         "background-color: transparent")
        self.execute_button.clicked.connect(self.execute_algorithm)
        left_layout.addWidget(self.execute_button)

        self.select_points_button = QPushButton(self)
        self.select_points_button.setStyleSheet("image: url(./grafika/but_klasy.png);"
                                         "width: 120px;"
                                         "height: 60px;"
                                         "margin: 0px;"
                                         "background-color: transparent")
        self.select_points_button.clicked.connect(self.show_select_points_dialog)
        left_layout.addWidget(self.select_points_button)


        self.back_button = QPushButton(self)
        self.back_button.setStyleSheet("image: url(./grafika/but_powrot.png);"
                                         "width: 120px;"
                                         "height: 60px;"
                                         "margin: 0px;"
                                         "background-color: transparent")
        self.back_button.clicked.connect(self.go_back)
        left_layout.addWidget(self.back_button)

        # Set left part width to 1/4th of the main window width

        # Right part (Matplotlib plot)
        self.matplotlib_widget = MatplotlibWidget(self.layout.parentWidget())  # Use the parent widget as the parent for MatplotlibWidget
        right_layout.addWidget(self.matplotlib_widget)  # Set right part width to 3/4th of the main window width

        self.kerfus_table = QTableView()

        stylesheet = "::section{Background-color:rgb(69, 67, 84);}"
        self.kerfus_table.horizontalHeader().setStyleSheet(stylesheet)

        stylesheet = "::section{Background-color:rgb(69, 67, 84);}"
        self.kerfus_table.verticalHeader().setStyleSheet(stylesheet)

        self.kerfus_table.setStyleSheet("background-color: rgb(69, 67, 84); color: white;border:2px;border-style: none;")
        right_layout.addWidget(self.kerfus_table)

        self.layout.addLayout(left_layout, 1)
        self.layout.addLayout(right_layout, 3)

        self.setLayout(self.layout)

        self.layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setAutoFillBackground(True)
        p = self.palette()
        p.setColor(self.backgroundRole(), QColor(69, 67, 84))  # You can set any color you want here
        self.setPalette(p)

    def go_back(self):
        self.parent().setCurrentIndex(0)

    def execute_algorithm(self):
        try:
            map_data = self.data_manager.get_data("map").copy()
            points_data = self.data_manager.get_data("points").copy()
            selection = self.data_manager.get_data("selection")
            if self.data_manager.get_data("is_user"):
                user_classes = self.data_manager.get_data("user_classes")
            else:
                user_classes = None


            shop = map_data.values.T
            arr = shop.copy()
            points = points_data.values

            points[:, 2] = -1 * points[:, 2]
            points[:, 3] = -1 * points[:, 3]

            points_ref = [(x, y) for x, y in points[:, :2]]

            kerfus = alg_RSM.run_rsm(shop, points, selection, points_ref, 9, user_classes)

            base_coords = np.argwhere(arr == 6)
            base_coords = tuple(base_coords[0])
            arr[base_coords] = 0

            kerfus = np.delete(kerfus, 0, axis=0)
            kerfus = np.delete(kerfus, 0, axis=1)

            kerfus2 = kerfus.copy()
            kerfus2[:, 3] = -1 * kerfus2[:, 3]
            kerfus2[:, 4] = -1 * kerfus2[:, 4]

            kerfus = np.append([[base_coords[0], base_coords[1], 0, 0, 0, 0, 0]], kerfus, axis=0)


            kerfus_tab = [["lp", "x", "y", "ci", "popularność", "szerokość przejazdu", "przeszkadzanie", "odległość"]]

            for i, el in enumerate(kerfus2):
                kerfus_tab = np.append(kerfus_tab, [np.append(i + 1, el)], axis=0)

            print(kerfus_tab)

            ax = self.matplotlib_widget.canvas.figure.add_subplot(111)

            shelves = np.argwhere(arr == 1)
            entrance = np.argwhere(arr == 2)
            exit = np.argwhere(arr == 3)
            bread = np.argwhere(arr == 4)
            meat = np.argwhere(arr == 5)

            ax.scatter(points[:, 0], points[:, 1], c='red', s=5)
            for ix in range(kerfus.shape[0] - 1):
                path = a_star.astar_search(arr, tuple(map(int, kerfus[ix, :2])), tuple(map(int, kerfus[ix + 1, :2])))
                path = np.array(path)
                ax.plot(path[:, 0], path[:, 1], '--', c='#1f77b4')
            ax.scatter(kerfus[:, 0], kerfus[:, 1])
            ax.scatter(shelves[:, 0], shelves[:, 1], c='lightblue', marker='s', s=26)
            ax.scatter(entrance[:, 0], entrance[:, 1], c='green', marker='s', s=26)
            ax.scatter(exit[:, 0], exit[:, 1], c='red', marker='s', s=26)
            ax.scatter(bread[:, 0], bread[:, 1], c='grey', marker='s', s=26)
            ax.scatter(meat[:, 0], meat[:, 1], c='orange', marker='s', s=26)

            for i in range(kerfus.shape[0]):
                ax.annotate(i, tuple(kerfus[i, :2]))

            ax.set_axisbelow(True)
            ax.xaxis.set_minor_locator(MultipleLocator(1))
            ax.yaxis.set_minor_locator(MultipleLocator(1))
            ax.set_aspect('equal', adjustable='box')
            ax.set_xlim((-0.5, 55.5))
            ax.set_ylim((28.5, -0.5))
            ax.grid(which='both', linewidth=0.25)
            ax.set_xticks([])
            ax.set_yticks([])

            self.matplotlib_widget.canvas.draw()

            model = KerfusTableModel(kerfus_tab)
            self.kerfus_table.setModel(model)
        except Exception as e:
            if str(e) == 'too many indices for array: array is 1-dimensional, but 2 were indexed':
                self.show_error_popup("Error in Algorithm", "Klasy są sprzeczne, wybierz inne lub użyj algorytmicznych")
            else:
                print("Error executing algorithm:", str(e))


    def show_error_popup(self, title, message):
        error_popup = QMessageBox()
        error_popup.setIcon(QMessageBox.Icon.Critical)
        error_popup.setWindowTitle(title)
        error_popup.setText(message)
        error_popup.exec()

    def show_select_points_dialog(self):
        select_points_dialog = SelectPointsDialog(self, self.data_manager)
        select_points_dialog.exec()


class SelectPointsDialog(QDialog):
    def __init__(self, parent, data_manager: DataManager):
        try:
            super().__init__(parent)

            self.data_manager = data_manager

            self.setWindowTitle("Select Points")

            self.layout = QVBoxLayout()
            self.selection = []

            self.algo_check = QCheckBox('Używaj klas Algorytmicznych', self)
            self.user_check = QCheckBox('Używaj klas Użytkownika', self)

            self.group = QButtonGroup(self)
            self.group.addButton(self.algo_check)
            self.group.addButton(self.user_check)

            self.layout.addWidget(self.algo_check)
            self.layout.addWidget(self.user_check)

            self.group.buttonClicked.connect(self.show_hide_widgets)

            self.table = QTableWidget(self)
            self.table.setColumnCount(7)
            self.table.setHorizontalHeaderLabels(["X", "Y", "wsp. Popularnosci", "szerokość alejki", "odleglosc od półki", "klasa 1", "klasa 2"])

            self.load_data_button = QPushButton("Load Data from Excel")
            self.load_data_button.clicked.connect(self.load_data_from_excel)

            self.check_A0 = QCheckBox("A0")
            self.check_A1 = QCheckBox("A1")
            self.check_A2 = QCheckBox("A2")
            self.check_A3 = QCheckBox("A3")

            self.group2 = QButtonGroup(self)
            self.group2.setExclusive(False)
            self.group2.addButton(self.check_A0)
            self.group2.addButton(self.check_A1)
            self.group2.addButton(self.check_A2)
            self.group2.addButton(self.check_A3)

            self.layout.addWidget(self.check_A0)
            self.layout.addWidget(self.check_A1)
            self.layout.addWidget(self.check_A2)
            self.layout.addWidget(self.check_A3)

            self.check_A0.setVisible(0)
            self.check_A1.setVisible(0)
            self.check_A2.setVisible(0)
            self.check_A3.setVisible(0)

            self.group2.buttonClicked.connect(self.show_hide_widgets)

            self.setLayout(self.layout)
            self.selected_points = {'class1': set(), 'class2': set()}

            self.layout.addWidget(self.load_data_button)
            self.layout.addWidget(self.table)

        except Exception as e:
            print("Error loading data:", str(e))

    def show_hide_widgets(self):
        try:
            algo = self.algo_check.isChecked()
            user = self.user_check.isChecked()

            self.table.setVisible(user)
            self.table.setRowCount(0)

            self.load_data_button.setVisible(user)

            self.check_A0.setVisible(algo)
            self.check_A1.setVisible(algo)
            self.check_A2.setVisible(algo)
            self.check_A3.setVisible(algo)

            selected_checkboxes = [button for button in self.group2.buttons() if button.isChecked()]
            if len(selected_checkboxes) > 2:
                # Uncheck the last selected checkbox
                selected_checkboxes[-1].setChecked(False)

            self.selection = [i for i, button in enumerate(self.group2.buttons()) if button.isChecked()]

            self.data_manager.set_data("selection", self.selection)
            self.data_manager.set_data("is_user", user)
            print(user)
        except Exception as e:
            print("Error loading data:", str(e))


    def load_data_from_excel(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")

        if file_path:
            # Read data from the Excel file using openpyxl
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook['punkty']

                # Skip the first row and start from the second column
                data = [
                    cell.value if (cell is not None and cell.value is not None) else ""
                    for row in sheet.iter_rows(min_row=2)
                    for cell in row[1:]
                ]

                self.populate_table(data, sheet.max_column - 1)  # Subtract 1 for skipping the first column
            except Exception as e:
                print("Error loading data:", str(e))

    def save_data(self):
        print(self.selected_points)

    def populate_table(self, data, max_column):
        # Clear existing table content
        self.table.setRowCount(0)

        # Populate the table with data
        row_count = len(data) // max_column
        remaining_items = len(data) % max_column

        if remaining_items > 0:
            row_count += 1  # Add an extra row for the remaining items

        for row in range(row_count):
            self.table.insertRow(row)
            for col in range(max_column):
                index = row * max_column + col
                if index < len(data):
                    value = data[index]
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row, col, item)

                    if col in (5, 6):  # Only for columns 6 and 7 (checkbox columns)
                        checkbox_item = QTableWidgetItem()
                        checkbox_item.setFlags(checkbox_item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                        checkbox_item.setCheckState(Qt.CheckState.Unchecked)
                        self.table.setItem(row, col, checkbox_item)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.table.itemChanged.connect(self.checkbox_state_changed)

    def checkbox_state_changed(self, item):
        # Ensure that only one checkbox is checked in each row
        if item.column() in (5, 6):
            other_col = 6 if item.column() == 5 else 5
            other_item = self.table.item(item.row(), other_col)

            # Remove the point from the other set if it's added to the current set
            if item.checkState() == Qt.CheckState.Checked:
                other_item.setCheckState(Qt.CheckState.Unchecked)
                class_name = f'class{item.column() - 4}'  # Convert column index to class name (1 or 2)
                other_class_name = f'class{other_item.column() - 4}'
                self.selected_points[other_class_name].discard(item.row())

            class_name = f'class{item.column() - 4}'  # Convert column index to class name (1 or 2)
            if item.checkState() == Qt.CheckState.Checked:
                self.selected_points[class_name].add(item.row())
            else:
                self.selected_points[class_name].discard(item.row())
            self.data_manager.set_data("user_classes", self.selected_points)